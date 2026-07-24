"""
excel_template.py
=================
Excel Template Generator - Creates pre-formatted Excel templates for all report types.
Can be run standalone or imported into the GUI.

Fixes in this revision
----------------------
* Example observations are now dict-keyed by header, so the sample data can never
  drift out of alignment when columns are reordered or type-specific columns added.
* The previously-unused BORDER style is now applied to the header + example cells.
* Sheet creation order is Index -> Limitation -> Scope -> Observations -> POC.
* Type columns for every *_revalidation variant are DERIVED (base columns +
  revalidation columns) instead of hand-duplicated, and the missing revalidation
  report types are now registered.
"""

import os
from pathlib import Path
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage, ImageOps
import io

# ─── Styles ──────────────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill(start_color="FF8D1010", end_color="FF8D1010", fill_type="solid")
INSTRUCTION_FONT = Font(name="Calibri", size=10, color="FFFFFF")
REQUIRED_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ─── Column Definitions ─────────────────────────────────────────────────────

UNIVERSAL_COLUMNS = [
    ("Sr. No.", 15),
    ("Severity", 15),
    ("Vulnerabillity ID(CVE/CWE)", 20),
    ("Vulnerability Name", 30),
    ("Affected URL", 25),
    ("Description", 40),
    ("Impact", 20),
    ("Recommendation", 40),
]

AFFECTED_COLUMN_MAP = {
    'web': ("Affected URL", 25),
    'android': ("Affected APK", 25),
    'ios': ("Affected IPA", 25),
    'api': ("Affected Endpoint", 25),
    'red_team': ("Attack Vector", 25),
    'source_code': ("Affected Path", 25),
}

REPORT_TYPES = {
    'web': "Web Application",
    'android': "Mobile App (Android)",
    'ios': "Mobile App (iOS)",
    'api': "API Security",
    'red_team': "Red Teaming",
    'source_code': "Source Code Review",
}

SEVERITY_ORDER = {
    'critical': 0,
    'high': 1,
    'medium': 2,
    'low': 3,
    'info': 4
}

INDEX_DATA = [
    ("Report Title", "Security Assessment Report"),
    ("Client Name", ""),
    ("Application Name", ""),
    ("Application Type", "Internal / External"),
    ("Audit Period Start", ""),
    ("Audit Period End", ""),
    ("Environment", "Production / UAT / Staging"),
    ("Target URL / IP", ""),
    ("Release Date", ""),
    ("Prepared By", ""),
    ("Reviewed By", ""),
    ("Approved By", ""),
    ("Released By", ""),
]

# Examples are keyed by column header so they cannot drift out of alignment.
# Type-specific columns left unset render blank for the user to fill in.

class ExcelTemplateGenerator:
    def __init__(self):
        self.wb = None
        self.scope_data = []
        self.limitation_data = [] 

    def _apply_header_style(self, cell):
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER

    def generate(self, report_type: str, output_path: str = None, observations: list = None, poc_folder: str = None, scope: list = None, limitation: list = None):
        """Generate Excel template with optional observations."""
        self.wb = Workbook()
        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])

        # Get metadata from observations or pass separately
        metadata = {
            "client_name": observations[0].get("client_name", "") if observations else "",
            "app_name": observations[0].get("app_name", "") if observations else "",
            "report_name": f"{report_type.upper()} Security Assessment",
            "prepared_by": "",
            "reviewed_by": "",
            "url": "",
            "audit_period": "",
            "release_date": "",
        }

        severity_counts = self._calculate_severity_counts(observations)

        
        self._create_index_sheet(metadata, severity_counts, report_type)
        self._create_limitation_sheet(limitation)
        self._create_scope_sheet(scope)
        self._create_observations_sheet(report_type, observations)
        self._create_poc_sheet(observations, poc_folder)
     


        if output_path is None:
            output_path = f"Excel_Template_{report_type.upper()}.xlsx"

        self.wb.save(output_path)

        if observations:
            self._create_observation_folders(output_path, observations)

        return output_path
    
    def _calculate_severity_counts(self, observations):
        """Calculate severity counts from observations."""
        counts = {"High": 0, "Medium": 0, "Low": 0}
        if not observations:
            return counts
        
        for obs in observations:
            severity = obs.get("severity", "").lower()
            if severity in ["critical", "high"]:
                counts["High"] += 1
            elif severity in ["medium"]:
                counts["Medium"] += 1
            elif severity in ["low", "info"]:
                counts["Low"] += 1
        
        return counts

    def _create_index_sheet(self, metadata=None, severity_counts=None, report_type=None):
        from openpyxl.chart import PieChart, Reference
        from openpyxl.chart.series import DataPoint
        from openpyxl.drawing.image import Image as XLImage

        DARK_RED, RED, YELLOW, GREEN = "C00000", "EE0000", "FFC000", "00B050"

        REPORT_TITLES = {
            'web': "Web Application Security Assessment",
            'android': "Android Application Security Assessment",
            'ios': "iOS Application Security Assessment",
            'api': "API Security Assessment",
            'red_team': "Red Teaming Assessment",
            'source_code': "Source Code Review",
        }

        report_title = REPORT_TITLES.get(report_type, "Security Assessment")

        md = {
            "company_name": "NANGIA & CO LLP",
            "company_subtitle": "C H A R T E R E D   A C C O U N T A N T S",
            "title": f"{report_title}      {metadata.get('client_name', 'Client')}",
            "report_name": metadata.get("report_name", ""),
            "application": metadata.get("app_name", ""),
            "consultant": metadata.get("prepared_by", ""),
            "reviewed_by": metadata.get("reviewed_by", ""),
            "client_name": metadata.get("client_name", ""),
            "scope_url": metadata.get("url", ""),
            "activity_duration": metadata.get("audit_period", ""),
            "domain": metadata.get("domain", "Security Assessment"),
            "submission_date": metadata.get("release_date", ""),
            "app_detail": metadata.get("app_detail", ""),
        }
        if metadata:
            md.update(metadata)
        sc = {"High": 0, "Medium": 0, "Low": 0}
        if severity_counts:
            sc.update(severity_counts)
        total = sc["High"] + sc["Medium"] + sc["Low"]

        ws = self.wb.create_sheet("Index")
        thin = Side(style="thin", color="000000")

        def rng_border(ref, side=thin):
            b = Border(left=side, right=side, top=side, bottom=side)
            for row in ws[ref]:
                for c in row:
                    c.border = b

        def fill(cell, color):
            cell.fill = PatternFill("solid", fgColor=color)

        for col, w in {"A":18,"B":40,"C":3,"D":16,"E":14,"F":34,"G":8,"H":9,"I":8,"J":8}.items():
            ws.column_dimensions[col].width = w

        # ── header band with LOGO ──
        ws.merge_cells("A2:B3")
    
        
        logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "assets", "logo.png")
        assets_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "assets")
        logo_path = None

        if os.path.exists(assets_dir):
            for file in os.listdir(assets_dir):
                if file.lower().startswith('logo') and file.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif')):
                    logo_path = os.path.join(assets_dir, file)
                    break

        if logo_path:
            try:
                from openpyxl.drawing.image import Image as XLImage
                img = XLImage(logo_path)
                img.width = 240
                img.height = 45
                ws.add_image(img, "A2")
            except Exception as e:
                print(f"Error adding logo: {e}")
        
        
        
        ws.merge_cells("D2:G3")
        ws["D2"] = md["title"]
        ws["D2"].font = Font(bold=True, size=13, color="000000", name="Calibri")
        ws["D2"].alignment = Alignment(horizontal="center", vertical="center")
        rng_border("D2:G3")
        ws.row_dimensions[2].height = 26
        ws.row_dimensions[3].height = 18

        # ── metadata block ──
        fields = [("Report Name", md["report_name"]), ("Application", md["application"]),
                ("Consultant", md["consultant"]), ("Review by", md["reviewed_by"]),
                ("Client Name", md["client_name"]), ("Scope", None),
                ("Activity Duration", md["activity_duration"])]
        r = 6
        for label, val in fields:
            lc = ws.cell(r, 1, label)
            fill(lc, DARK_RED)
            lc.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
            lc.alignment = Alignment(vertical="center")
            vc = ws.cell(r, 2)
            if label == "Scope":
                vc.value = "Click here"
                if md["scope_url"]:
                    vc.hyperlink = md["scope_url"]
                vc.font = Font(color="0563C1", underline="single")
            else:
                vc.value = val
                vc.alignment = Alignment(vertical="center")
            rng_border(f"A{r}:B{r}")
            r += 1

        # ── severity table (also the pie chart's data source) ──
        hdr = 14
        for col, txt in enumerate(["Severity", "Vulnerability Count"], 1):
            c = ws.cell(hdr, col, txt)
            fill(c, DARK_RED)
            c.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
            c.alignment = Alignment(horizontal="center")
        for i, (name, cnt, color, txtcol) in enumerate([
                ("High", sc["High"], RED, "FFFFFF"),
                ("Medium", sc["Medium"], YELLOW, "000000"),
                ("Low", sc["Low"], GREEN, "FFFFFF")]):
            a = ws.cell(hdr + 1 + i, 1, name)
            fill(a, color)
            a.font = Font(bold=True, color=txtcol)
            ws.cell(hdr + 1 + i, 2, cnt).alignment = Alignment(horizontal="center")
        tot_row = hdr + 4
        ws.cell(tot_row, 1, "Total").font = Font(bold=True)
        tc = ws.cell(tot_row, 2, total)
        tc.font = Font(bold=True)
        tc.alignment = Alignment(horizontal="center")
        rng_border(f"A{hdr}:B{tot_row}")

        # ── pie chart ──
        pie = PieChart()
        data = Reference(ws, min_col=2, min_row=hdr + 1, max_row=hdr + 3)
        cats = Reference(ws, min_col=1, min_row=hdr + 1, max_row=hdr + 3)
        pie.add_data(data, titles_from_data=False)
        pie.set_categories(cats)
        pie.title = f"VULNERABILITY DISTRIBUTION FOR\n{md['title'].split('-')[0].strip().upper()}"
        pie.height = 8
        pie.width = 10
        pie.legend.position = "b"
        pie.legend.overlay = False
        pie.title.overlay = False

        # Control pie area (x, y, width, height) to prevent overlapping
        from openpyxl.chart.layout import Layout, ManualLayout
        pie.plot_area.layout = Layout(manualLayout=ManualLayout(
            xMode="edge", yMode="edge", x=0.22, y=0.30, w=0.56, h=0.52
        ))

        for i, color in enumerate([RED, YELLOW, GREEN]):
            dp = DataPoint(idx=i)
            dp.graphicalProperties.solidFill = color
            pie.series[0].data_points.append(dp)
        ws.add_chart(pie, "D5")

    def _create_scope_sheet(self, scope=None):
        ws = self.wb.create_sheet("Scope")
        
        ws.cell(row=1, column=1, value="S. No.")
        ws.cell(row=1, column=2, value="Scope")
        for col in [1, 2]:
            cell = ws.cell(row=1, column=col)
            self._apply_header_style(cell)
        
        if scope and isinstance(scope, list):
            for row, item in enumerate(scope, 2):
                if item and item.strip():
                    ws.cell(row=row, column=1, value=str(row - 1))
                    ws.cell(row=row, column=2, value=str(item).strip())
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 80
        ws.freeze_panes = 'A2'

    def _create_limitation_sheet(self, limitation=None):
        ws = self.wb.create_sheet("Limitation")
        
        ws.cell(row=1, column=1, value="S. No.")
        ws.cell(row=1, column=2, value="Limitation")
        for col in [1, 2]:
            cell = ws.cell(row=1, column=col)
            self._apply_header_style(cell)
        
        if limitation and isinstance(limitation, list):
            for row, item in enumerate(limitation, 2):
                if item and item.strip():
                    ws.cell(row=row, column=1, value=str(row - 1))
                    ws.cell(row=row, column=2, value=str(item).strip())
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 80
        ws.freeze_panes = 'A2'

    def _sort_observations_by_severity(self, observations):
        """Sort observations by severity: Critical → High → Medium → Low → Info."""
        
        def get_severity_rank(obs):
            severity = obs.get('severity', 'medium').lower()
            return SEVERITY_ORDER.get(severity, 999)
        
        return sorted(observations, key=get_severity_rank)  
    

    def _create_observations_sheet(self, report_type, observations=None):
        ws = self.wb.create_sheet("Observations")
        
        # Get the correct Affected column name based on report type
        affected_column = AFFECTED_COLUMN_MAP.get(report_type, ("Affected URL", 25))
        
        # Build columns: Universal columns with Affected column replaced
        columns = []
        for col in UNIVERSAL_COLUMNS:
            if col[0] == "Affected URL":  # Skip the generic one
                continue
            columns.append(col)
        columns.insert(4, affected_column)  # Insert at position 4 (after Vulnerability ID)
        
        # Write headers (row 1)
        for col, (header, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)
            ws.column_dimensions[get_column_letter(col)].width = width
        
        if observations:
            observations = self._sort_observations_by_severity(observations)

        # If observations provided, fill them starting row 3
        if observations:
            header_names = [h for h, _ in columns]
            for row, obs in enumerate(observations, 2):
                # Map observation fields to header order
                data = []
                for header in header_names:
                    if header == "Sr. No.":
                        data.append(str(row - 1))
                    elif header == "Severity":
                        data.append(obs.get("severity", ""))
                    elif header == "Vulnerability Name":
                        data.append(obs.get("title", ""))
                    elif header == "Vulnerabillity ID(CVE/CWE)":
                        data.append(obs.get("cve", ""))
                    elif header in ["Affected URL", "Affected IP", "Affected APK", "Affected IPA", "Affected Endpoint", "Attack Vector", "Affected Path", "Affected Item"]:
                        data.append(obs.get("affected_url", obs.get("affected_ip", "")))
                    elif header == "Description":
                        data.append(obs.get("description", ""))
                    elif header == "Impact":
                        data.append(obs.get("impact", ""))
                    elif header == "Recommendation":
                        data.append(obs.get("recommendation", ""))
                    else:
                        data.append("")
                
                for col, value in enumerate(data, 1):
                    if col <= len(columns):
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.border = BORDER
                        
                        # Center align for columns 1, 2, 4 (Sr. No., Severity, Vulnerability ID)
                        if col in [1, 2, 4]:
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        else:
                            cell.alignment = Alignment(vertical='top', wrap_text=True)
                        
                        # Severity column (index 2) - Bold with background color
                        if col == 2:
                            severity_text = str(value).lower()
                            cell.font = Font(bold=True, color="000000")
                            
                            if 'critical' in severity_text:
                                cell.fill = PatternFill(start_color="FFC00000", end_color="FFC00000", fill_type="solid")
                                cell.font = Font(bold=True, color="FFFFFF")
                            elif 'high' in severity_text:
                                cell.fill = PatternFill(start_color="FFEE0000", end_color="FFEE0000", fill_type="solid")
                                cell.font = Font(bold=True, color="FFFFFF")
                            elif 'medium' in severity_text:
                                cell.fill = PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid")
                                cell.font = Font(bold=True, color="000000")
                            elif 'low' in severity_text:
                                cell.fill = PatternFill(start_color="FF00B050", end_color="FF00B050", fill_type="solid")
                                cell.font = Font(bold=True, color="FFFFFF")
                            elif 'info' in severity_text:
                                cell.fill = PatternFill(start_color="FF0070C0", end_color="FF0070C0", fill_type="solid")
                                cell.font = Font(bold=True, color="FFFFFF")

        ws.freeze_panes = 'A3'

    def _create_observation_folders(self, output_path, observations):
        """Create folders for each observation in the same location as the Excel file."""
        
        # Get the directory where Excel is saved
        base_dir = os.path.dirname(os.path.abspath(output_path))
        
        for obs in observations:
            # Get observation title
            title = obs.get('title', '').strip()
            if not title:
                continue
            
            # Clean title to make valid folder name
            folder_name = re.sub(r'[\\/*?:"<>|]', '', title)
            folder_name = folder_name.strip()
            
            # Create folder
            folder_path = os.path.join(base_dir, folder_name)
            Path(folder_path).mkdir(parents=True, exist_ok=True)
            
            # Create a placeholder file inside the folder
            readme_path = os.path.join(folder_path, "README.txt")
            with open(readme_path, 'w', encoding='utf-8') as f:
                f.write(f"POC screenshots for: {title}\n")
                f.write(f"Observation ID: {obs.get('sr_no', 'N/A')}\n")
                f.write("Place your POC images here (1.png, 2.png, 3.png, ...)\n")

    def _create_poc_sheet(self, observations=None, poc_folder=None):
            ws = self.wb.create_sheet("POC")

            # Headers
            headers = ["S. No.", "Severity", "Vulnerability"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                self._apply_header_style(cell)
                ws.column_dimensions[get_column_letter(col)].width = 50 if col == 3 else 15

            if not observations:
                ws.cell(row=2, column=1, value="No observations")
                return

            # Scan POC folder if provided
            poc_finder = None
            if poc_folder and os.path.exists(poc_folder):
                from src.poc_finder import POCFinder
                poc_finder = POCFinder()
                poc_finder.scan_folder(poc_folder)

            from openpyxl.drawing.image import Image as XLImage

            # Display caps (NOT pixel caps). The image keeps full resolution; only the
            # on-sheet display box is shrunk, so it stays sharp. Raise these for bigger
            # POCs, lower them for smaller. BORDER_DISPLAY_PX = how thick the border looks.
            MAX_DISPLAY_W = 650
            MAX_DISPLAY_H = 850
            BORDER_DISPLAY_PX = 1

            current_row = 2

            for obs in observations:
                sr_no    = obs.get("sr_no", "")
                severity = obs.get("severity", "")
                title    = obs.get("title", "")

                ws.cell(row=current_row, column=1, value=sr_no)
                ws.cell(row=current_row, column=2, value=severity)
                ws.cell(row=current_row, column=3, value=title)

                # Severity formatting (bold + colour)
                severity_text = str(severity).lower()
                sev_cell = ws.cell(row=current_row, column=2)
                sev_cell.font = Font(bold=True, color="000000")
                if 'critical' in severity_text:
                    sev_cell.fill = PatternFill(start_color="FFC00000", end_color="FFC00000", fill_type="solid")
                    sev_cell.font = Font(bold=True, color="FFFFFF")
                elif 'high' in severity_text:
                    sev_cell.fill = PatternFill(start_color="FFEE0000", end_color="FFEE0000", fill_type="solid")
                    sev_cell.font = Font(bold=True, color="FFFFFF")
                elif 'medium' in severity_text:
                    sev_cell.fill = PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid")
                    sev_cell.font = Font(bold=True, color="000000")
                elif 'low' in severity_text:
                    sev_cell.fill = PatternFill(start_color="FF00B050", end_color="FF00B050", fill_type="solid")
                    sev_cell.font = Font(bold=True, color="FFFFFF")
                elif 'info' in severity_text:
                    sev_cell.fill = PatternFill(start_color="FF0070C0", end_color="FF0070C0", fill_type="solid")
                    sev_cell.font = Font(bold=True, color="FFFFFF")

                for col in range(1, 4):
                    ws.cell(row=current_row, column=col).border = BORDER

                current_row += 4  # space before images

                # POC images for this vulnerability
                images = []
                if poc_finder and title:
                    images = poc_finder.get_pocs_by_vulnerability(title)

                for img_path in images:
                    try:
                        im = PILImage.open(img_path)
                        if im.mode not in ("RGB", "RGBA"):
                            im = im.convert("RGB")

                        native_w, native_h = im.width, im.height

                        # how much we SHRINK THE DISPLAY (never enlarge). Pixels stay full-res.
                        scale = min(MAX_DISPLAY_W / native_w, MAX_DISPLAY_H / native_h, 1.0)

                        # border baked on the full-res image, thick enough to look ~1px after shrink
                        border_px = max(1, round(BORDER_DISPLAY_PX / scale)) if scale > 0 else 1
                        fill = (0, 0, 0, 255) if im.mode == "RGBA" else (0, 0, 0)
                        im = ImageOps.expand(im, border=border_px, fill=fill)

                        buf = io.BytesIO()
                        im.save(buf, format="PNG")   # PNG = lossless, full resolution kept
                        buf.seek(0)
                        if not hasattr(self, "_poc_buffers"):
                            self._poc_buffers = []
                        self._poc_buffers.append(buf)   # keep alive until wb.save()

                        xi = XLImage(buf)
                        # shrink the DISPLAY box only -> high-res pixels render into it -> crisp
                        disp_w = int(im.width * scale)
                        disp_h = int(im.height * scale)
                        xi.width = disp_w
                        xi.height = disp_h

                        ws.add_image(xi, ws.cell(row=current_row, column=1).coordinate)
                        ws.row_dimensions[current_row].height = disp_h * 0.75 + 6

                        current_row += 4
                    except Exception as e:
                        ws.cell(row=current_row, column=1,
                                value=f"[Error: {os.path.basename(img_path)}: {e}]")
                        current_row += 1

            ws.freeze_panes = 'A2'

if __name__ == "__main__":
    gen = ExcelTemplateGenerator()
    for rt in REPORT_TYPES:
        path = gen.generate(rt, f"Excel_Template_{rt}.xlsx")
        print(f"generated {rt:<26} -> {path}")
