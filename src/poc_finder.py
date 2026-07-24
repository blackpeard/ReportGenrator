"""
POC Finder - Matches POC references with actual files
Supports both reference-based and folder-based matching
"""
from pathlib import Path
import glob
import os
import re

class POCFinder:
    def __init__(self):
        self.poc_files = {}
        self.poc_folder = None
        self.folder_structure = {}  # New: Store folder-based POCs
        
    def scan_folder(self, folder_path):
        """Scan folder and index all image files"""
        self.poc_folder = Path(folder_path)
        self.poc_files = {}
        self.folder_structure = {}
        
        image_extensions = ['*.png', '*.jpg', '*.jpeg', '*.gif', '*.bmp']
        
        for ext in image_extensions:
            for file in self.poc_folder.rglob(ext):
                # Original indexing for reference-based matching
                name = file.stem.lower()
                full_name = file.name.lower()
                
                self.poc_files[name] = str(file)
                self.poc_files[full_name] = str(file)
                
                # Clean name indexing
                clean_name = re.sub(r'[^a-zA-Z0-9]', '', name)
                if clean_name != name:
                    self.poc_files[clean_name] = str(file)
                
                # NEW: Index by folder structure (vulnerability name)
                parent_folder = file.parent.name
                if parent_folder not in self.folder_structure:
                    self.folder_structure[parent_folder] = []
                
                self.folder_structure[parent_folder].append(str(file))
        
        # Sort images in each folder naturally
        for folder in self.folder_structure:
            self.folder_structure[folder].sort(
                key=lambda x: [int(c) if c.isdigit() else c for c in re.split(r'(\d+)', x)]
            )
        
        return len(self.poc_files)
    
    def find_poc(self, reference):
        """Find POC file matching the reference (original method)"""
        if not reference or not self.poc_files:
            return None
            
        ref = str(reference).strip().lower()
        
        # Direct match
        if ref in self.poc_files:
            return self.poc_files[ref]
        
        # Try without extension
        ref_without_ext = Path(ref).stem
        if ref_without_ext in self.poc_files:
            return self.poc_files[ref_without_ext]
        
        # Try partial match
        for key, path in self.poc_files.items():
            if ref in key or key in ref:
                return path
        
        return None  
 
    # NEW METHOD 1: Get POCs by vulnerability title (folder name)
    def get_pocs_by_vulnerability(self, vulnerability_title):
        """
        Get all POC images for a vulnerability based on folder name
        
        Args:
            vulnerability_title: Title of the vulnerability from Excel
            
        Returns:
            List of image paths sorted naturally, or empty list if not found
        """
        if not vulnerability_title or not self.folder_structure:
            return []
        
        # Clean the title to match folder name
        clean_title = re.sub(r'[\\/*?:"<>|]', '', vulnerability_title).strip()
        
        # Try exact match first
        if clean_title in self.folder_structure:
            return self.folder_structure[clean_title]
        
        # Try case-insensitive match
        for folder in self.folder_structure:
            if folder.lower() == clean_title.lower():
                return self.folder_structure[folder]
        
        # Try partial match (if folder name contains title or vice versa)
        for folder, images in self.folder_structure.items():
            if (clean_title.lower() in folder.lower()) or (folder.lower() in clean_title.lower()):
                return images
        
        return []
    
    def load_pocs_from_excel(self, excel_path: str):
        """
        Load POC data from Excel POC sheet and extract embedded images AND steps.
        Preserves the exact order as they appear in Excel.
        """
        import pandas as pd
        from openpyxl import load_workbook
        from PIL import Image
        import io
        import tempfile
        import os
        import re
        
        poc_sheet_names = ["POC", "POCs", "Proof of Concept", "Screenshots"]
        
        wb = None
        sheet_name_found = None
        
        for sheet_name in poc_sheet_names:
            try:
                wb = load_workbook(excel_path, data_only=True)
                if sheet_name in wb.sheetnames:
                    sheet_name_found = sheet_name
                    print(f"  ✓ Found POC sheet: '{sheet_name}'")
                    break
            except:
                continue
        
        if wb is None or sheet_name_found is None:
            print("  ⚠ No POC sheet found in Excel")
            self.excel_poc_items = {}
            return False
        
        # Read the sheet with pandas for text data
        df = pd.read_excel(excel_path, sheet_name=sheet_name_found)
        df.columns = [str(col).strip() for col in df.columns]
        
        # Detect columns
        col_mapping = {}
        for col in df.columns:
            col_lower = str(col).strip().lower()
            if 'sr' in col_lower or 's.no' in col_lower or '#' in col_lower:
                col_mapping['sr_no'] = col
            elif 'severity' in col_lower:
                col_mapping['severity'] = col
            elif 'vulnerability' in col_lower or 'title' in col_lower or 'observation' in col_lower:
                col_mapping['vulnerability'] = col
        
        # Get the worksheet for embedded images
        ws = wb[sheet_name_found]
        
        # Extract embedded images with their row numbers
        images_by_row = {}
        for img in ws._images:
            if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                row = img.anchor._from.row + 1
                if row not in images_by_row:
                    images_by_row[row] = []
                images_by_row[row].append(img)
        
        self.excel_poc_items = {}
        temp_dir = tempfile.gettempdir()
        
        # Get all observation rows
        obs_rows = []
        for idx, row in df.iterrows():
            excel_row = idx + 2
            if 'vulnerability' in col_mapping:
                val = row.get(col_mapping['vulnerability'], "")
                if pd.notna(val) and str(val).strip():
                    obs_rows.append((excel_row, str(val).strip()))
        
        # Process each observation
        for obs_idx, (obs_row, vulnerability) in enumerate(obs_rows):
            # Determine the next observation row
            if obs_idx + 1 < len(obs_rows):
                next_obs_row = obs_rows[obs_idx + 1][0]
            else:
                next_obs_row = 9999
            
            # Collect ALL items in the order they appear in Excel
            ordered_items = []
            image_counter = 1
            
            for row_num in range(obs_row + 1, next_obs_row):
                # Check if this row has text (steps)
                step_text = None
                for col_idx in range(1, len(df.columns) + 1):
                    try:
                        cell_value = ws.cell(row=row_num, column=col_idx).value
                        if cell_value and isinstance(cell_value, str):
                            text = str(cell_value).strip()
                            if text and not text.startswith('S. No.') and not text.startswith('Severity'):
                                if text.strip():
                                    step_text = text
                                    break
                    except:
                        pass
                
                if step_text:
                    ordered_items.append(('step', step_text))
                
                # Check if this row has images
                if row_num in images_by_row:
                    for img in images_by_row[row_num]:
                        ordered_items.append(('image', img, image_counter))
                        image_counter += 1
            
            # Extract and save images, keep steps as text
            items_for_this_obs = []
            
            for item in ordered_items:
                if item[0] == 'step':
                    items_for_this_obs.append(('step', item[1]))
                elif item[0] == 'image':
                    img = item[1]
                    img_idx = item[2]
                    try:
                        img_data = None
                        if hasattr(img, '_data') and callable(img._data):
                            img_data = img._data()
                        elif hasattr(img, 'image'):
                            if hasattr(img.image, 'fp'):
                                img.image.fp.seek(0)
                                img_data = img.image.fp.read()
                            elif hasattr(img.image, 'getvalue'):
                                img_data = img.image.getvalue()
                            elif isinstance(img.image, (bytes, bytearray)):
                                img_data = img.image
                        elif hasattr(img, 'ref') and isinstance(img.ref, (bytes, bytearray)):
                            img_data = img.ref
                        
                        if not img_data:
                            continue
                        
                        clean_title = re.sub(r'[\\/*?:"<>|]', '', vulnerability).strip()
                        temp_path = os.path.join(temp_dir, f"{clean_title}_{img_idx}.png")
                        img_obj = Image.open(io.BytesIO(img_data))
                        img_obj.save(temp_path)
                        items_for_this_obs.append(('image', temp_path))
                        
                    except Exception as e:
                        print(f"    ⚠️  Could not extract image: {e}")
            
            if items_for_this_obs:
                clean_title = re.sub(r'[\\/*?:"<>|]', '', vulnerability).strip()
                self.excel_poc_items[clean_title] = items_for_this_obs
                print(f"  ✓ Extracted {len(items_for_this_obs)} items for: {vulnerability[:40]}")
        
        wb.close()
        print(f"  ✓ Found POCs for {len(self.excel_poc_items)} vulnerabilities")
        return True
        
    def get_excel_poc_items_by_vulnerability(self, vulnerability_title):
        """
        Get POC items (steps and images in order) from Excel data.
        Returns list of ('step', text) or ('image', path) in the order they appear.
        """
        if not hasattr(self, 'excel_poc_items') or not self.excel_poc_items:
            return []
        
        if not vulnerability_title:
            return []
        
        clean_title = re.sub(r'[\\/*?:"<>|]', '', vulnerability_title).strip()
        
        if clean_title in self.excel_poc_items:
            return self.excel_poc_items[clean_title]
        
        for title, items in self.excel_poc_items.items():
            if title.lower() == clean_title.lower():
                return items
        
        return []
    
    def get_excel_steps_by_vulnerability(self, vulnerability_title):
        """
        Get POC steps from Excel data by vulnerability title.
        """
        if not hasattr(self, 'excel_poc_steps') or not self.excel_poc_steps:
            return []
        
        if not vulnerability_title:
            return []
        
        clean_title = re.sub(r'[\\/*?:"<>|]', '', vulnerability_title).strip()
        
        if clean_title in self.excel_poc_steps:
            return self.excel_poc_steps[clean_title]
        
        for title, steps in self.excel_poc_steps.items():
            if title.lower() == clean_title.lower():
                return steps
        
        return []
    
    # NEW METHOD 2: Check if vulnerability has POCs
    def has_pocs(self, vulnerability_title):
        """Check if a vulnerability has any POC images"""
        return len(self.get_pocs_by_vulnerability(vulnerability_title)) > 0
    
    # NEW METHOD 3: Get folder names (for debugging/listing)
    def get_all_vulnerability_folders(self):
        """Get list of all folder names that have POCs"""
        return list(self.folder_structure.keys())
